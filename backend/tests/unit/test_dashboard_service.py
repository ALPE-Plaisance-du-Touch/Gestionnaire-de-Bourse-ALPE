"""Unit tests for dashboard and export services."""

import pytest
from datetime import datetime, timezone, timedelta
from decimal import Decimal
from unittest.mock import MagicMock

from app.models import User
from app.models.article import Article, ArticleStatus
from app.models.item_list import ItemList
from app.models.payout import Payout, PayoutStatus


# ---- Helpers ----

def _make_user(**kwargs):
    user = MagicMock(spec=User)
    user.id = kwargs.get("id", "user-1")
    user.email = kwargs.get("email", "test@example.com")
    user.first_name = kwargs.get("first_name", "Jean")
    user.last_name = kwargs.get("last_name", "Dupont")
    user.is_active = kwargs.get("is_active", False)
    user.password_hash = kwargs.get("password_hash", None)
    user.invitation_token = kwargs.get("invitation_token", "token-123")
    user.invitation_expires_at = kwargs.get("invitation_expires_at", datetime.now(timezone.utc) + timedelta(days=7))
    user.invitation_hidden = kwargs.get("invitation_hidden", False)
    user.role_id = kwargs.get("role_id", 1)
    user.created_at = kwargs.get("created_at", datetime.now(timezone.utc) - timedelta(days=10))
    user.updated_at = kwargs.get("updated_at", datetime.now(timezone.utc) - timedelta(days=9))
    return user


def _make_payout(**kwargs):
    payout = MagicMock(spec=Payout)
    payout.id = kwargs.get("id", "payout-1")
    payout.gross_amount = kwargs.get("gross_amount", Decimal("10.00"))
    payout.commission_amount = kwargs.get("commission_amount", Decimal("2.00"))
    payout.list_fees = kwargs.get("list_fees", Decimal("0.00"))
    payout.net_amount = kwargs.get("net_amount", Decimal("8.00"))
    payout.total_articles = kwargs.get("total_articles", 5)
    payout.sold_articles = kwargs.get("sold_articles", 3)
    payout.unsold_articles = kwargs.get("unsold_articles", 2)
    payout.status = kwargs.get("status", PayoutStatus.PENDING.value)
    payout.payment_method = kwargs.get("payment_method", None)
    payout.depositor_id = kwargs.get("depositor_id", "user-1")
    payout.notes = kwargs.get("notes", None)

    depositor = _make_user(id=payout.depositor_id)
    payout.depositor = depositor

    item_list = MagicMock(spec=ItemList)
    item_list.number = kwargs.get("list_number", 100)
    item_list.list_type = kwargs.get("list_type", "standard")

    articles = []
    for i in range(payout.sold_articles):
        a = MagicMock(spec=Article)
        a.line_number = i + 1
        a.status = ArticleStatus.SOLD.value
        a.description = f"Article vendu {i+1}"
        a.category = "clothing"
        a.price = Decimal("3.00")
        a.size = None
        a.brand = None
        articles.append(a)
    for i in range(payout.unsold_articles):
        a = MagicMock(spec=Article)
        a.line_number = payout.sold_articles + i + 1
        a.status = ArticleStatus.ON_SALE.value
        a.description = f"Article invendu {i+1}"
        a.category = "toys"
        a.price = Decimal("5.00")
        a.size = None
        a.brand = None
        articles.append(a)

    item_list.articles = articles
    payout.item_list = item_list

    return payout


# ---- Payout Excel Tests ----

class TestPayoutExcel:

    def test_generate_4_sheets(self):
        from app.services.payout_excel import generate_payout_excel
        from openpyxl import load_workbook
        from io import BytesIO

        edition = MagicMock()
        edition.name = "Bourse Automne 2025"

        payouts = [_make_payout(), _make_payout(id="payout-2", list_number=101)]
        stats = {
            "total_sales": Decimal("20.00"),
            "total_commission": Decimal("4.00"),
            "total_list_fees": Decimal("0.00"),
            "total_net": Decimal("16.00"),
            "total_articles": 10,
            "sold_articles": 6,
            "unsold_articles": 4,
            "total_payouts": 2,
            "payouts_pending": 2,
            "payouts_ready": 0,
            "payouts_paid": 0,
            "payouts_cancelled": 0,
        }

        result = generate_payout_excel(payouts, edition, stats)
        assert isinstance(result, bytes)
        assert len(result) > 0

        wb = load_workbook(BytesIO(result))
        assert len(wb.sheetnames) == 4
        assert wb.sheetnames[0] == "Recapitulatif deposants"
        assert wb.sheetnames[1] == "Detail ventes"
        assert wb.sheetnames[2] == "Invendus"
        assert wb.sheetnames[3] == "Statistiques"

    def test_depositor_summary_rows(self):
        from app.services.payout_excel import generate_payout_excel
        from openpyxl import load_workbook
        from io import BytesIO

        edition = MagicMock()
        edition.name = "Test"
        payouts = [_make_payout(), _make_payout(id="p2", list_number=200)]
        stats = {"total_sales": Decimal("0"), "total_commission": Decimal("0"),
                 "total_list_fees": Decimal("0"), "total_net": Decimal("0"),
                 "total_articles": 0, "sold_articles": 0, "unsold_articles": 0,
                 "total_payouts": 2, "payouts_pending": 2, "payouts_ready": 0,
                 "payouts_paid": 0, "payouts_cancelled": 0}

        wb = load_workbook(BytesIO(generate_payout_excel(payouts, edition, stats)))
        ws = wb["Recapitulatif deposants"]
        # Header + 2 data rows + 1 totals row
        assert ws.max_row == 4

    def test_sold_articles_sheet(self):
        from app.services.payout_excel import generate_payout_excel
        from openpyxl import load_workbook
        from io import BytesIO

        edition = MagicMock()
        edition.name = "Test"
        payout = _make_payout(sold_articles=4, unsold_articles=1, total_articles=5)
        stats = {"total_sales": Decimal("0"), "total_commission": Decimal("0"),
                 "total_list_fees": Decimal("0"), "total_net": Decimal("0"),
                 "total_articles": 0, "sold_articles": 0, "unsold_articles": 0,
                 "total_payouts": 1, "payouts_pending": 1, "payouts_ready": 0,
                 "payouts_paid": 0, "payouts_cancelled": 0}

        wb = load_workbook(BytesIO(generate_payout_excel([payout], edition, stats)))
        ws = wb["Detail ventes"]
        # Header + 4 sold articles
        assert ws.max_row == 5


# ---- Invitation Excel Tests ----

class TestInvitationExcel:

    def test_generate_3_sheets(self):
        from app.services.invitation_excel import generate_invitation_excel
        from openpyxl import load_workbook
        from io import BytesIO

        users = [_make_user(), _make_user(id="u2", email="b@test.com")]
        stats = {"total": 2, "activated": 0, "pending": 2, "expired": 0,
                 "activation_rate": 0.0, "avg_activation_delay_days": 0.0,
                 "expiration_rate": 0.0, "relaunch_count": 0,
                 "activated_after_relaunch": 0}

        result = generate_invitation_excel(users, stats)
        assert isinstance(result, bytes)

        wb = load_workbook(BytesIO(result))
        assert len(wb.sheetnames) == 3
        assert wb.sheetnames[0] == "Liste invitations"
        assert wb.sheetnames[1] == "Statistiques"
        assert wb.sheetnames[2] == "A relancer"

    def test_to_relaunch_filter(self):
        from app.services.invitation_excel import generate_invitation_excel
        from openpyxl import load_workbook
        from io import BytesIO

        # 1 pending, 1 activated
        pending = _make_user(id="u1", is_active=False, invitation_token="tok")
        activated = _make_user(
            id="u2", is_active=True, password_hash="hashed",
            invitation_token=None,
        )
        stats = {"total": 2, "activated": 1, "pending": 1, "expired": 0,
                 "activation_rate": 50.0, "avg_activation_delay_days": 1.0,
                 "expiration_rate": 0.0, "relaunch_count": 0,
                 "activated_after_relaunch": 0}

        wb = load_workbook(BytesIO(generate_invitation_excel([pending, activated], stats)))
        ws = wb["A relancer"]
        # Header + 1 pending user only
        assert ws.max_row == 2


# ---- Invitation Stats Tests ----

class TestInvitationStats:

    def test_compute_status_activated(self):
        from app.services.invitation import InvitationService
        user = _make_user(is_active=True, password_hash="hashed")
        now = datetime.now(timezone.utc)
        assert InvitationService._compute_status(user, now) == "activated"

    def test_compute_status_pending(self):
        from app.services.invitation import InvitationService
        user = _make_user(
            is_active=False, invitation_token="tok",
            invitation_expires_at=datetime.now(timezone.utc) + timedelta(days=3),
        )
        now = datetime.now(timezone.utc)
        assert InvitationService._compute_status(user, now) == "pending"

    def test_compute_status_expired(self):
        from app.services.invitation import InvitationService
        user = _make_user(
            is_active=False, invitation_token="tok",
            invitation_expires_at=datetime.now(timezone.utc) - timedelta(days=1),
        )
        now = datetime.now(timezone.utc)
        assert InvitationService._compute_status(user, now) == "expired"


# ---- Closure Report PDF Tests ----

class TestClosureReportPDF:

    def test_generate_returns_bytes(self):
        from app.services.closure_report_pdf import generate_closure_report_pdf

        edition = MagicMock()
        edition.name = "Bourse Test"
        edition.location = "Salle des fetes"

        payout = _make_payout()
        stats = {
            "total_sales": Decimal("100.00"),
            "total_commission": Decimal("20.00"),
            "total_list_fees": Decimal("1.00"),
            "total_net": Decimal("79.00"),
            "total_articles": 10,
            "sold_articles": 7,
            "unsold_articles": 3,
            "total_payouts": 1,
            "payouts_paid": 0,
        }

        result = generate_closure_report_pdf(edition, stats, [payout], "Admin Test")
        assert isinstance(result, bytes)
        assert len(result) > 100  # Non-trivial PDF

    def test_empty_payouts(self):
        from app.services.closure_report_pdf import generate_closure_report_pdf

        edition = MagicMock()
        edition.name = "Empty"
        edition.location = None

        stats = {
            "total_sales": Decimal("0"), "total_commission": Decimal("0"),
            "total_list_fees": Decimal("0"), "total_net": Decimal("0"),
            "total_articles": 0, "sold_articles": 0, "unsold_articles": 0,
            "total_payouts": 0, "payouts_paid": 0,
        }

        result = generate_closure_report_pdf(edition, stats, [])
        assert isinstance(result, bytes)
        assert len(result) > 0
