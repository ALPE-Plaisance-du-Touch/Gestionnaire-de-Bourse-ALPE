"""Excel export service for payout data."""

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.article import ArticleStatus

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CURRENCY_FORMAT = '#,##0.00 "EUR"'


def _apply_header_style(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT


def _auto_column_width(ws):
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted = min(max_length + 3, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def _add_auto_filter(ws):
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def generate_payout_excel(payouts, edition, stats: dict) -> bytes:
    """Generate Excel file with 4 sheets for payout export."""
    wb = Workbook()

    _build_depositor_summary(wb, payouts)
    _build_sales_detail(wb, payouts)
    _build_unsold_detail(wb, payouts)
    _build_global_stats(wb, edition, stats, payouts)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_depositor_summary(wb, payouts):
    ws = wb.active
    ws.title = "Recapitulatif deposants"

    headers = [
        "N liste", "Deposant", "Type liste", "Articles total",
        "Vendus", "Invendus", "Ventes brut", "Commission",
        "Frais liste", "Net a reverser", "Statut", "Methode paiement",
    ]
    ws.append(headers)
    _apply_header_style(ws)

    for p in payouts:
        depositor = p.depositor
        ws.append([
            p.item_list.number,
            f"{depositor.first_name} {depositor.last_name}",
            p.item_list.list_type,
            p.total_articles,
            p.sold_articles,
            p.unsold_articles,
            float(p.gross_amount),
            float(p.commission_amount),
            float(p.list_fees),
            float(p.net_amount),
            p.status,
            p.payment_method or "",
        ])

    # Apply currency format to financial columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=10):
        for cell in row:
            cell.number_format = CURRENCY_FORMAT

    # SUM formulas at bottom
    if len(payouts) > 0:
        last_data_row = len(payouts) + 1
        sum_row = last_data_row + 1
        ws.cell(row=sum_row, column=1, value="TOTAUX").font = Font(bold=True)
        for col in (4, 5, 6, 7, 8, 9, 10):
            col_letter = get_column_letter(col)
            cell = ws.cell(
                row=sum_row, column=col,
                value=f"=SUM({col_letter}2:{col_letter}{last_data_row})",
            )
            cell.font = Font(bold=True)
            if col >= 7:
                cell.number_format = CURRENCY_FORMAT

    _auto_column_width(ws)
    _add_auto_filter(ws)


def _build_sales_detail(wb, payouts):
    ws = wb.create_sheet("Detail ventes")

    headers = [
        "N liste", "Deposant", "N article", "Description",
        "Categorie", "Prix", "Taille", "Marque",
    ]
    ws.append(headers)
    _apply_header_style(ws)

    for p in payouts:
        depositor = p.depositor
        depositor_name = f"{depositor.first_name} {depositor.last_name}"
        for article in sorted(p.item_list.articles, key=lambda a: a.line_number):
            if article.status == ArticleStatus.SOLD.value:
                ws.append([
                    p.item_list.number,
                    depositor_name,
                    article.line_number,
                    article.description,
                    article.category,
                    float(article.price),
                    article.size or "",
                    article.brand or "",
                ])

    # Currency format for price column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = CURRENCY_FORMAT

    _auto_column_width(ws)
    _add_auto_filter(ws)


def _build_unsold_detail(wb, payouts):
    ws = wb.create_sheet("Invendus")

    headers = [
        "N liste", "Deposant", "N article", "Description",
        "Categorie", "Prix demande", "Taille", "Marque",
    ]
    ws.append(headers)
    _apply_header_style(ws)

    unsold_statuses = (
        ArticleStatus.ON_SALE.value,
        ArticleStatus.UNSOLD.value,
        ArticleStatus.RETRIEVED.value,
        ArticleStatus.DONATED.value,
    )

    for p in payouts:
        depositor = p.depositor
        depositor_name = f"{depositor.first_name} {depositor.last_name}"
        for article in sorted(p.item_list.articles, key=lambda a: a.line_number):
            if article.status in unsold_statuses:
                ws.append([
                    p.item_list.number,
                    depositor_name,
                    article.line_number,
                    article.description,
                    article.category,
                    float(article.price),
                    article.size or "",
                    article.brand or "",
                ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.number_format = CURRENCY_FORMAT

    _auto_column_width(ws)
    _add_auto_filter(ws)


def _build_global_stats(wb, edition, stats: dict, payouts):
    ws = wb.create_sheet("Statistiques")

    label_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14, color="2563EB")

    ws.cell(row=1, column=1, value="STATISTIQUES GLOBALES").font = title_font
    ws.cell(row=2, column=1, value=f"Edition : {edition.name}")
    ws.merge_cells("A1:B1")

    data = [
        ("", ""),
        ("Resultats financiers", ""),
        ("Total des ventes", stats.get("total_sales", Decimal("0.00"))),
        ("Commission ALPE", stats.get("total_commission", Decimal("0.00"))),
        ("Frais de liste", stats.get("total_list_fees", Decimal("0.00"))),
        ("Net a reverser", stats.get("total_net", Decimal("0.00"))),
        ("", ""),
        ("Resultats articles", ""),
        ("Total articles", stats.get("total_articles", 0)),
        ("Articles vendus", stats.get("sold_articles", 0)),
        ("Articles invendus", stats.get("unsold_articles", 0)),
        ("", ""),
        ("Reversements", ""),
        ("Total reversements", stats.get("total_payouts", 0)),
        ("En attente", stats.get("payouts_pending", 0)),
        ("Prets", stats.get("payouts_ready", 0)),
        ("Payes", stats.get("payouts_paid", 0)),
        ("Annules", stats.get("payouts_cancelled", 0)),
    ]

    for row_idx, (label, value) in enumerate(data, start=4):
        cell_label = ws.cell(row=row_idx, column=1, value=label)
        cell_value = ws.cell(row=row_idx, column=2, value=float(value) if isinstance(value, Decimal) else value)
        if label and not value and value != 0:
            cell_label.font = Font(bold=True, size=12, color="2563EB")
        else:
            cell_label.font = label_font
        if isinstance(value, Decimal):
            cell_value.number_format = CURRENCY_FORMAT

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
