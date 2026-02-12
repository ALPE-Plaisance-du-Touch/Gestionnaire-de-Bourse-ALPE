"""Excel export service for invitation data."""

from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_header_style(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT


def _auto_column_width(ws):
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted = min(max_length + 3, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def _add_auto_filter(ws):
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def _compute_status(user, now) -> str:
    if user.is_active and user.password_hash:
        return "Activee"
    if not user.invitation_token:
        return "Annulee"
    if user.invitation_expires_at:
        expires = user.invitation_expires_at
        if expires.tzinfo is None:
            expires = expires.replace(tzinfo=timezone.utc)
        if now > expires:
            return "Expiree"
    return "En attente"


def generate_invitation_excel(users: list, stats: dict) -> bytes:
    """Generate Excel file with 3 sheets for invitation export."""
    wb = Workbook()
    now = datetime.now(timezone.utc)

    _build_full_list(wb, users, now)
    _build_statistics(wb, stats)
    _build_to_relaunch(wb, users, now)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_full_list(wb, users, now):
    ws = wb.active
    ws.title = "Liste invitations"

    headers = [
        "Email", "Prenom", "Nom", "Statut",
        "Date creation", "Date expiration", "Date activation",
    ]
    ws.append(headers)
    _apply_header_style(ws)

    for user in users:
        status = _compute_status(user, now)
        activation_date = ""
        if status == "Activee" and user.updated_at:
            activation_date = user.updated_at.strftime("%Y-%m-%d %H:%M")

        ws.append([
            user.email,
            user.first_name or "",
            user.last_name or "",
            status,
            user.created_at.strftime("%Y-%m-%d %H:%M") if user.created_at else "",
            user.invitation_expires_at.strftime("%Y-%m-%d %H:%M") if user.invitation_expires_at else "",
            activation_date,
        ])

    _auto_column_width(ws)
    _add_auto_filter(ws)


def _build_statistics(wb, stats: dict):
    ws = wb.create_sheet("Statistiques")

    title_font = Font(bold=True, size=14, color="2563EB")
    label_font = Font(bold=True, size=11)

    ws.cell(row=1, column=1, value="STATISTIQUES INVITATIONS").font = title_font
    ws.merge_cells("A1:B1")

    data = [
        ("Total invitations", stats.get("total", 0)),
        ("Activees", stats.get("activated", 0)),
        ("En attente", stats.get("pending", 0)),
        ("Expirees", stats.get("expired", 0)),
        ("", ""),
        ("Taux d'activation (%)", stats.get("activation_rate", 0.0)),
        ("Delai moyen d'activation (jours)", stats.get("avg_activation_delay_days", 0.0)),
        ("Taux d'expiration (%)", stats.get("expiration_rate", 0.0)),
        ("Nombre de relances", stats.get("relaunch_count", 0)),
        ("Activees apres relance", stats.get("activated_after_relaunch", 0)),
    ]

    for row_idx, (label, value) in enumerate(data, start=3):
        ws.cell(row=row_idx, column=1, value=label).font = label_font
        ws.cell(row=row_idx, column=2, value=value)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15


def _build_to_relaunch(wb, users, now):
    ws = wb.create_sheet("A relancer")

    headers = ["Email", "Prenom", "Nom", "Statut", "Date creation", "Date expiration"]
    ws.append(headers)
    _apply_header_style(ws)

    for user in users:
        status = _compute_status(user, now)
        if status in ("En attente", "Expiree"):
            ws.append([
                user.email,
                user.first_name or "",
                user.last_name or "",
                status,
                user.created_at.strftime("%Y-%m-%d %H:%M") if user.created_at else "",
                user.invitation_expires_at.strftime("%Y-%m-%d %H:%M") if user.invitation_expires_at else "",
            ])

    _auto_column_width(ws)
    _add_auto_filter(ws)
